from BoardInteractor import BoardInteractor
import openpyxl
from openpyxl import *
from datetime import datetime
import os

# This class will be used to read data from the BoardInteractor, And store it to an excel file.
class DataHandler:
    fileNamePrefix = "Gas_Measurment_"
    def __init__(self,boardInteractor):
        self.interactor = boardInteractor

    def getDayMonthYear(self):
        return datetime.now().day+"_"+self.getMonthYear()
    
    def getMonthYear(self):
        return datetime.now().month+"_"+datetime.now().year

    def getFileName(self):
        return self.fileNamePrefix + datetime.now().month+"_"+datetime.now().year

    def getFileNameWithSuffix(self):
        return self.getFileName()+".xlsx"

    def createSheet(self,wb:Workbook):
        sheet = wb.create_sheet(self.getDayMonthYear())
        return sheet

    def searchForWorkbook(self):

        if not os.path.exists(self.getFileNameWithSuffix()):
            print("Workbook not found, creating work book.")
            self.setupWorkbook()
        else:
            print("Workbook found.")
            try:
                workbook = openpyxl.load_workbook(self.getFileNameWithSuffix())
                print(f"Sheets found: {workbook.sheetnames}")

                #Want to get the latest sheet, but we also want to see if the latest sheet is today. 
                if self.getDayMonthYear() in workbook.sheetnames:
                    print(f"Sheet with for date (name) {self.getDayMonthYear()} was found.")
                    sheet = workbook[self.getDayMonthYear()]
                else:
                    print(f"Sheet with for date (name) {self.getDayMonthYear()} was NOT found.")
                    print(f"Sheets found: {workbook.sheetnames}")
                    sheet = self.createSheet(workbook)


            except Exception as E:
                print("Filed to load workbook from existing file. "+E.with_traceback())

    def setupWorkbook(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Pressure Measurement Data "+self.getDayMonthYear()
        sheet = self.createSheet(workbook)
        workbook.save(self.getFileNameWithSuffix())
        
    